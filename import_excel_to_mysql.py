"""
从 Excel 读取6家门店预估客流/业绩数据，写入 MySQL。
Sheet 名称即日期，日期与当天不匹配时发送钉钉通知。
缺少门店数据表时发送钉钉通知。
"""
import os
import sys
import re
import json
import time
import hmac
import hashlib
import base64
import urllib.parse
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl
import mysql.connector
import requests

# ============================================================
# 配置 — 按需修改
# ============================================================
MYSQL_CONFIG = {
    "host": "192.168.32.132",
    "port": 3306,
    "user": "root",
    "password": "123456a",
    "database": "门店小程序销售数据",
    "charset": "utf8mb4",
}

DINGTALK_ACCESS_TOKEN = "eda7e7c6a2072b350775cf3fbf5f8ca0e9e96cebd4e4f45cdae39984a3277ec3"
DINGTALK_SECRET = "SECd65fd7ea8c3c0d04b8fe41def1ea22dfb409cdfe2d2c2c46493904659f8382f5"
DINGTALK_BASE_URL = "https://oapi.dingtalk.com/robot/send"

TABLE_NAME = "预计客流业绩"
POINT_TABLE_NAME = "点位人效"
STORE_DIR_TABLE = "门店目录"
POINT_MAPPING_TABLE = "点位映射配置"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)


# ============================================================
# 日期解析
# ============================================================
def parse_date(raw: str) -> date | None:
    """尝试解析各种非标准日期格式，返回 date 对象或 None"""
    if not raw or not isinstance(raw, str):
        return None

    raw = raw.strip()

    # 已经是 datetime 对象
    if isinstance(raw, datetime):
        return raw.date()

    # ISO 格式: 2026-05-01 / 2026/05/01 / 2026.05.01
    for sep in ("-", "/", "."):
        try:
            return datetime.strptime(raw, f"%Y{sep}%m{sep}%d").date()
        except ValueError:
            pass

    # 无分隔符: 20260501
    if re.match(r"^\d{8}$", raw):
        try:
            return datetime.strptime(raw, "%Y%m%d").date()
        except ValueError:
            pass

    # 中文格式: 2026年05月01日 / 2026年5月1日
    m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", raw)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    # 月日格式: 5月1日 / 05-01  (补齐当前年份)
    m = re.match(r"(\d{1,2})月(\d{1,2})日", raw)
    if m:
        return date(date.today().year, int(m.group(1)), int(m.group(2)))

    m = re.match(r"^(\d{1,2})-(\d{1,2})$", raw)
    if m:
        return date(date.today().year, int(m.group(1)), int(m.group(2)))

    log.warning(f"无法解析日期: {raw!r}")
    return None


# ============================================================
# Excel 数据提取
# ============================================================
def extract_data(filepath: str) -> dict | None:
    """从 Excel 提取门店名称、预估客流、预估业绩、日期"""
    filepath = str(filepath)
    if not os.path.exists(filepath):
        log.error(f"文件不存在: {filepath}")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 最后一个 sheet 名称即为日期
    sheet_name = wb.sheetnames[-1]
    sheet_date = parse_date(sheet_name)
    if sheet_date is None:
        log.error(f"Sheet 名称无法解析为日期: {sheet_name!r}")
        wb.close()
        return None

    ws = wb[sheet_name]

    # 第1行第1列 = 门店名称
    store_name = ws.cell(row=1, column=1).value
    if store_name and isinstance(store_name, datetime):
        store_name = store_name.strftime("%Y-%m-%d")
    store_name = str(store_name).strip() if store_name else "未知"

    # 在第2行中搜索 "预估客流" 和 "预估业绩"
    header_row = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val:
            header_row[str(val).strip()] = col

    footfall_col = header_row.get("预估客流")
    sales_col = header_row.get("预估业绩")

    if footfall_col is None:
        log.error(f"未找到'预估客流'列，表头: {list(header_row.keys())}")
        wb.close()
        return None
    if sales_col is None:
        log.error(f"未找到'预估业绩'列，表头: {list(header_row.keys())}")
        wb.close()
        return None

    # 第3行取数据
    footfall_raw = ws.cell(row=3, column=footfall_col).value
    sales_raw = ws.cell(row=3, column=sales_col).value

    try:
        footfall = int(footfall_raw) if footfall_raw else 0
    except (ValueError, TypeError):
        footfall = 0
    try:
        sales = int(sales_raw) if sales_raw else 0
    except (ValueError, TypeError):
        sales = 0

    wb.close()

    result = {
        "门店名称": store_name,
        "日期": sheet_date,
        "预估客流": footfall,
        "预估业绩": sales,
        "sheet_name_raw": sheet_name,
    }
    log.info(f"提取结果: {result}")
    return result


# ============================================================
# 点位数据提取
# ============================================================
def extract_points(filepath: str) -> dict:
    """
    从 Excel 提取点位名称和对应的出勤人次。
    返回: {Excel点位名: 人数}
    """
    filepath = str(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]

    # 找到"点位 | 出勤人次"表头行
    header_row = None
    for row in range(1, ws.max_row + 1):
        v1 = ws.cell(row=row, column=1).value
        v2 = ws.cell(row=row, column=2).value
        if v1 and v2:
            if str(v1).strip() == "点位" and "出勤人次" in str(v2):
                header_row = row
                break

    if header_row is None:
        log.warning(f"未找到'点位 | 出勤人次'表头: {filepath}")
        wb.close()
        return {}

    # 数据从 header_row + 2 开始（跳过子表头行）
    points = {}
    for row in range(header_row + 2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        count = ws.cell(row=row, column=2).value
        if name and count is not None:
            name = str(name).strip()
            try:
                count = int(count)
            except (ValueError, TypeError):
                continue
            points[name] = count
        # 空 name 时继续（子行/员工行），直到遇到下个有 name 的行

    wb.close()
    log.info(f"点位提取: {points}")
    return points


def load_point_mapping(store_name: str) -> dict:
    """从数据库 点位映射配置 表加载某门店的点位映射。返回 {DB点位名: [Excel原始点位, ...]}"""
    try:
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        cursor = conn.cursor()
        cursor.execute(
            f"SELECT `数据库点位`, `Excel原始点位` FROM `{POINT_MAPPING_TABLE}` WHERE `门店` = %s",
            (store_name,),
        )
        mapping = {}
        for db_name, excel_name in cursor.fetchall():
            mapping.setdefault(db_name, []).append(excel_name)
        cursor.close()
        conn.close()
        return mapping
    except Exception as e:
        log.error(f"加载点位映射失败 ({store_name}): {e}")
        return {}


def apply_point_mapping(excel_points: dict, store_name: str) -> dict:
    """
    从数据库加载映射，将 Excel 点位名合并为数据库点位名。
    返回: {DB点位名: 在岗人数}
    """
    mapping = load_point_mapping(store_name)
    if not mapping:
        log.warning(f"门店无点位映射配置: {store_name}")
        return {}

    result = {}
    for db_name, excel_names in mapping.items():
        total = sum(excel_points.get(en, 0) for en in excel_names)
        result[db_name] = total
    return result


# ============================================================
# MySQL 操作
# ============================================================
def upsert_data(data: dict) -> bool:
    """写入数据（按 门店+日期 唯一约束，重复则更新）"""
    try:
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        cursor = conn.cursor()

        sql = f"""
            INSERT INTO `{TABLE_NAME}` (`门店`, `日期`, `预估客流`, `业绩目标`)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                `预估客流` = VALUES(`预估客流`),
                `业绩目标` = VALUES(`业绩目标`)
        """
        cursor.execute(sql, (data["门店名称"], data["日期"], data["预估客流"], data["预估业绩"]))
        conn.commit()

        log.info(f"写入成功: 门店={data['门店名称']}, 日期={data['日期']}, "
                 f"客流={data['预估客流']}, 业绩={data['预估业绩']}")
        return True
    except Exception as e:
        log.error(f"MySQL 写入失败: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ============================================================
# 点位数据写入
# ============================================================
def upsert_point_data(store_name: str, record_date: date, db_points: dict) -> int:
    """
    写入点位人效数据（按 日期+门店+点位 唯一约束，重复则更新今日在岗人数）。
    返回写入条数。
    """
    if not db_points:
        return 0

    try:
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        cursor = conn.cursor()

        sql = f"""
            INSERT INTO `{POINT_TABLE_NAME}` (`门店`, `日期`, `点位`, `今日在岗人数`)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                `今日在岗人数` = VALUES(`今日在岗人数`)
        """
        count = 0
        for point_name, headcount in db_points.items():
            cursor.execute(sql, (store_name, record_date, point_name, headcount))
            count += 1

        conn.commit()
        log.info(f"点位写入成功: {store_name} | {record_date} | {count} 个点位")
        return count
    except Exception as e:
        log.error(f"点位写入失败: {e}")
        return 0
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ============================================================
# 钉钉通知（加签模式）
# ============================================================
def _dingtalk_sign() -> str:
    """生成钉钉加签参数: ?timestamp=xxx&sign=xxx"""
    timestamp = str(round(time.time() * 1000))
    string_to_sign = f"{timestamp}\n{DINGTALK_SECRET}"
    hmac_code = hmac.new(
        DINGTALK_SECRET.encode("utf-8"),
        string_to_sign.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).digest()
    sign = urllib.parse.quote_plus(base64.b64encode(hmac_code).decode("utf-8"))
    return f"access_token={DINGTALK_ACCESS_TOKEN}&timestamp={timestamp}&sign={sign}"


def _dingtalk_post(text: str, title: str):
    """发送钉钉 markdown 消息"""
    payload = {
        "msgtype": "markdown",
        "markdown": {"title": title, "text": text},
    }
    url = f"{DINGTALK_BASE_URL}?{_dingtalk_sign()}"
    try:
        resp = requests.post(url, json=payload, timeout=10)
        body = resp.json()
        if body.get("errcode") == 0:
            log.info(f"钉钉通知已发送: {title}")
        else:
            log.warning(f"钉钉通知发送失败: {body}")
    except Exception as e:
        log.error(f"钉钉通知异常: {e}")


def send_date_mismatch_alert(data: dict, today: date):
    """日期不匹配时通知"""
    text = (
        f"## Excel 日期不匹配告警\n\n"
        f"- 门店: **{data['门店名称']}**\n"
        f"- 文件日期: **{data['sheet_name_raw']}**\n"
        f"- 当前日期: **{today}**\n"
        f"- 预估客流: {data['预估客流']}\n"
        f"- 预估业绩: {data['预估业绩']}\n\n"
        f"> 数据未写入数据库，请核实后重新导入。"
    )
    _dingtalk_post(text, "Excel 日期不匹配告警")


def send_missing_stores_alert(missing: list, found: list):
    """缺少门店数据表时通知"""
    store_lines = "\n".join(f"- **{s}**" for s in missing)
    found_lines = "\n".join(f"- {s}" for s in found) if found else "- 无"
    text = (
        f"## 缺少门店数据表\n\n"
        f"### 缺少的门店（共 {len(missing)} 家）：\n"
        f"{store_lines}\n\n"
        f"### 已有的门店（共 {len(found)} 家）：\n"
        f"{found_lines}\n\n"
        f"> 请补充缺失门店的 Excel 文件后重新导入。"
    )
    _dingtalk_post(text, "缺少门店数据表")


# ============================================================
# 门店列表
# ============================================================
def load_standard_stores() -> list:
    """从数据库 门店目录 表动态读取标准门店名称"""
    try:
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        cursor = conn.cursor()
        cursor.execute(f"SELECT `门店` FROM `{STORE_DIR_TABLE}` ORDER BY `nc_order`")
        stores = [r[0] for r in cursor.fetchall()]
        cursor.close()
        conn.close()
        log.info(f"从 {STORE_DIR_TABLE} 读取到 {len(stores)} 家门店: {stores}")
        return stores
    except Exception as e:
        log.error(f"读取门店目录失败: {e}")
        return []


# ============================================================
# 文件发现
# ============================================================
def find_excel_files(directory: str) -> list:
    """在目录中查找 Excel 文件"""
    files = []
    for f in os.listdir(directory):
        if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$"):
            files.append(os.path.join(directory, f))
    return files


def match_files_to_stores(filepaths: list, standard_stores: list) -> dict:
    """
    将 Excel 文件匹配到标准门店。
    返回 dict: {门店名称: filepath}，同时跳过无法匹配的文件。
    """
    matched = {}
    for fp in filepaths:
        data = extract_data(fp)
        if data is None:
            log.warning(f"无法提取数据，跳过: {fp}")
            continue
        name = data["门店名称"]
        if name in standard_stores:
            matched[name] = fp
        else:
            log.warning(f"门店名称不在标准列表中，跳过: {name!r} (文件: {fp})")
    return matched


# ============================================================
# 主流程
# ============================================================
def main():
    # 固定读取 NAS 目录
    data_dir = "/vol1/1000/rizhi"

    if len(sys.argv) >= 2:
        # 兼容旧用法：传入单个文件
        excel_files = [sys.argv[1]]
    else:
        excel_files = find_excel_files(data_dir)
        if not excel_files:
            print(f"[错误] 目录下未找到 Excel 文件: {data_dir}")
            sys.exit(1)

    today = date.today()

    # 0. 从数据库加载门店列表
    STANDARD_STORES = load_standard_stores()
    if not STANDARD_STORES:
        print("[错误] 无法从数据库加载门店列表，退出。")
        sys.exit(1)

    # 1. 匹配文件到标准门店
    matched = match_files_to_stores(excel_files, STANDARD_STORES)
    found_stores = list(matched.keys())
    missing_stores = [s for s in STANDARD_STORES if s not in matched]

    # 2. 缺少门店 → 钉钉通知
    if missing_stores:
        log.warning(f"缺少门店数据: {missing_stores}")
        send_missing_stores_alert(missing_stores, found_stores)
        print(f"\n[警告] 缺少 {len(missing_stores)} 家门店数据: {', '.join(missing_stores)}")
    else:
        print(f"[OK] 6 家门店数据表齐全")

    if not found_stores:
        print("[错误] 没有任何有效门店数据，退出。")
        sys.exit(1)

    # 3. 逐店处理
    success_list = []
    point_success_list = []
    fail_list = []
    mismatch_list = []

    for store_name, filepath in matched.items():
        data = extract_data(filepath)
        if data is None:
            fail_list.append(store_name)
            continue

        # 日期校验
        if data["日期"] != today:
            log.warning(f"[{store_name}] 日期不匹配: {data['日期']} != {today}")
            send_date_mismatch_alert(data, today)
            mismatch_list.append((store_name, data['sheet_name_raw']))
            continue

        # 写入客流/业绩
        if upsert_data(data):
            success_list.append((store_name, data["预估客流"], data["预估业绩"]))
        else:
            fail_list.append(store_name)

        # 提取并写入点位数据
        excel_points = extract_points(filepath)
        if excel_points:
            db_points = apply_point_mapping(excel_points, store_name)
            if db_points:
                count = upsert_point_data(store_name, data["日期"], db_points)
                if count:
                    point_success_list.append((store_name, count, db_points))
                else:
                    log.warning(f"[{store_name}] 点位写入失败")
            else:
                log.warning(f"[{store_name}] 点位映射后为空")
        else:
            log.warning(f"[{store_name}] 未提取到点位数据")

    # 4. 汇总输出
    print(f"\n{'='*50}")
    print(f"  处理日期: {today}")
    print(f"  客流/业绩导入: {len(success_list)} 家")
    for s in success_list:
        print(f"    [OK] {s[0]}  客流={s[1]}  业绩={s[2]}")
    print(f"  点位人效导入: {len(point_success_list)} 家")
    for p in point_success_list:
        detail = ", ".join(f"{k}={v}" for k, v in p[2].items())
        print(f"    [OK] {p[0]}  ({p[1]}个点位) {detail}")
    if mismatch_list:
        print(f"  日期不匹配: {len(mismatch_list)} 家")
        for m in mismatch_list:
            print(f"    [跳过] {m[0]} (文件日期={m[1]})")
    if fail_list:
        print(f"  失败: {len(fail_list)} 家")
        for f in fail_list:
            print(f"    [失败] {f}")
    if missing_stores:
        print(f"  缺表: {len(missing_stores)} 家")
        for m in missing_stores:
            print(f"    [缺表] {m}")
    print(f"{'='*50}\n")


if __name__ == "__main__":
    main()
